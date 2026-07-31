#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📥 SISTEMA DE RELATÓRIOS - PDF e EXCEL
Gera relatórios semanais, mensais e históricos
"""

import json
from datetime import datetime, timedelta
from io import BytesIO
import csv

def gerar_relatorio_excel(chamados, periodo='mensal'):
    """Gerar relatório em Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Chamados"
        
        # Header
        ws['A1'] = "RELATÓRIO DE CHAMADOS - UNIFEB"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Período: {periodo.upper()}"
        ws['A2'].font = Font(bold=True)
        
        # Colunas
        colunas = ['ID', 'Título', 'Solicitante', 'Prioridade', 'Status', 'Setor', 'Data Abertura', 'SLA']
        for col, titulo in enumerate(colunas, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = titulo
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        # Dados
        for row, chamado in enumerate(chamados, 5):
            ws.cell(row=row, column=1).value = chamado.get('id')
            ws.cell(row=row, column=2).value = chamado.get('titulo')
            ws.cell(row=row, column=3).value = chamado.get('solicitante')
            ws.cell(row=row, column=4).value = chamado.get('prioridade')
            ws.cell(row=row, column=5).value = chamado.get('status')
            ws.cell(row=row, column=6).value = chamado.get('setorAtendimento')
            
            # Formatar data/hora corretamente
            data_str = chamado.get('data', '')
            try:
                from datetime import datetime
                data_obj = datetime.fromisoformat(data_str.replace('Z', '+00:00'))
                data_formatada = data_obj.strftime('%d/%m/%Y, %H:%M:%S')
                ws.cell(row=row, column=7).value = data_formatada
            except:
                ws.cell(row=row, column=7).value = data_str
            
            # Cor por SLA
            sla_cell = ws.cell(row=row, column=8)
            if chamado.get('slaVencido'):
                sla_cell.value = "🔴 VENCIDO"
                sla_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif chamado.get('status') == 'Aberto':
                sla_cell.value = "🟡 EM PRAZO"
                sla_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                sla_cell.value = "🟢 OK"
                sla_cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        
        # Ajustar largura
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        
        # Salvar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
        
    except ImportError:
        print("❌ openpyxl não instalado. Use: pip install openpyxl")
        return None

def gerar_relatorio_csv(chamados):
    """Gerar relatório em CSV"""
    try:
        output = BytesIO()
        writer = csv.DictWriter(output, fieldnames=[
            'id', 'titulo', 'solicitante', 'prioridade', 'status',
            'setorAtendimento', 'data', 'slaVencido'
        ])
        
        output = BytesIO()
        writer_text = csv.writer(output.detach() if hasattr(output, 'detach') else output)
        
        # Header
        writer_text.writerow([
            'ID', 'Título', 'Solicitante', 'Prioridade', 'Status',
            'Setor', 'Data', 'SLA Vencido'
        ])
        
        # Dados
        for chamado in chamados:
            writer_text.writerow([
                chamado.get('id'),
                chamado.get('titulo'),
                chamado.get('solicitante'),
                chamado.get('prioridade'),
                chamado.get('status'),
                chamado.get('setorAtendimento'),
                chamado.get('data'),
                'Sim' if chamado.get('slaVencido') else 'Não'
            ])
        
        return output
        
    except Exception as e:
        print(f"❌ Erro ao gerar CSV: {e}")
        return None

def calcular_metricas(chamados):
    """Calcular métricas para dashboard executivo"""
    total = len(chamados)
    vencidos = sum(1 for c in chamados if c.get('slaVencido'))
    abertos = sum(1 for c in chamados if c.get('status') == 'Aberto')
    resolvidos = sum(1 for c in chamados if c.get('status') == 'Resolvido')
    
    por_prioridade = {
        'Alta': sum(1 for c in chamados if c.get('prioridade') == 'Alta'),
        'Média': sum(1 for c in chamados if c.get('prioridade') == 'Média'),
        'Baixa': sum(1 for c in chamados if c.get('prioridade') == 'Baixa')
    }
    
    por_status = {
        'Aberto': abertos,
        'Em andamento': sum(1 for c in chamados if c.get('status') == 'Em andamento'),
        'Resolvido': resolvidos
    }
    
    return {
        'total': total,
        'vencidos': vencidos,
        'percentual_vencidos': (vencidos / total * 100) if total > 0 else 0,
        'abertos': abertos,
        'resolvidos': resolvidos,
        'por_prioridade': por_prioridade,
        'por_status': por_status,
        'tempo_medio_resolucao': '~24h',  # Será calculado depois
        'saude_geral': 'Crítica' if (vencidos / total * 100) > 50 else 'Boa' if (vencidos / total * 100) < 20 else 'Alerta'
    }

if __name__ == "__main__":
    # Teste
    with open('chamados_sync.json', 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    chamados = dados.get('chamados', [])
    metricas = calcular_metricas(chamados)
    
    print("📊 MÉTRICAS:")
    print(f"  Total: {metricas['total']}")
    print(f"  Vencidos: {metricas['vencidos']} ({metricas['percentual_vencidos']:.1f}%)")
    print(f"  Abertos: {metricas['abertos']}")
    print(f"  Resolvidos: {metricas['resolvidos']}")
    print(f"  Saúde: {metricas['saude_geral']}")
